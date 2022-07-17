###written by Znydee and Rafe
import smtplib
import csv
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
#import pandas as pd
import re
import json
import openpyxl
regex = '^[a-z0-9]+[\._]?[a-z0-9]+[@]\w+[.]\w{2,3}$' 
def check(email):
    if(re.fullmatch(pattern = regex, string = email)):
        return "Valid Email"
    else:
        return "Invalid Email"
def submit_list():
    while True:
        try:
            email_list = json.loads(input("Please, paste your list here:  " ))
            break
        except:
            print("Please, Enter a valid list")
    not_valid_emails=[]
    for recv in email_list:          
        email_status=check(recv)
        if email_status=="Valid Email":
            msg['To'] = recv
            smtp_server.sendmail(Sender_Email,recv,msg.as_string())
        else:
            not_valid_emails.append(recv)
    print(f"successfully sent to {len(email_list)-len(not_valid_emails)} of {len(email_list)}")
    if len(not_valid_emails) == 0:
        pass
    else:
        print(f"not sent to {not_valid_emails}")
    smtp_server.quit()

def adding_manually():
    email_list = []
    while True:
        email_input = input("To: ")
        email_status=check(email_input)
        print(email_status)
        if email_status=="Valid Email":
            email_list.append(email_input)
            print("Enter '+' to add emails or 's' to send")
            opt=input(">>>")
            if opt=="+":
                pass
            elif opt=="s":
                break
            else:
                print("Invalid option")
        else:
            print("Please enter a valid email ")
    not_valid_emails=[]
    for recv in email_list:          
        email_status=check(recv)
        if email_status=="Valid Email":
            msg['To'] = recv
            smtp_server.sendmail(Sender_Email,recv,msg.as_string())
        else:
            not_valid_emails.append(recv)
    print(f"successfully sent to {len(email_list)-len(not_valid_emails)} of {len(email_list)}")
    if len(not_valid_emails) == 0:
        pass
    else:
        print(f"not sent to {not_valid_emails}")
    smtp_server.quit()

def excel():
    while True:
        pat=input('Enter the path address of the File: ')
        try:
            e = openpyxl.load_workbook(pat)
            break
        except:
            print('Invalid Address')
    while True:
        sh_name=input("enter the name of the sheet: ")
        try:
            sheet=e[sh_name]            
            break
        except:
            print('enter correct sheet name')
    while True:
        col= input('Enter the name of the Column: ')        
        i=1          
        while i<= sheet.max_column:
            if col == sheet.cell(1,i).value:         
                col=i
                col_present="yes"
                break
            else:      
                i=i+1 
                col_present="no"            
        if col_present=="yes":
            break
        else:
            print('Invalid Column')
    print("processing...")
    email_list = []
    i=1
    while i <= sheet.max_row:
        email_list.append(sheet.cell(i,col).value) 
        i=i+1               
    not_valid_emails=[]    
    print("sending now")
    for recv in email_list:          
        email_status=check(recv)
        if email_status=="Valid Email":
            msg['To'] = recv
            smtp_server.sendmail(Sender_Email,recv,msg.as_string())
        else:
            not_valid_emails.append(recv)
    print(f"successfully sent to {len(email_list)-len(not_valid_emails)} of {len(email_list)}")
    if len(not_valid_emails) == 0:
        pass
    else:
        print(f"not sent to {not_valid_emails}")
    smtp_server.quit()

def csv_send():
    rows=[]
    while True:
        pat=input('Enter the path address of the File: ')
        try:
            with open(pat, 'r') as csvfile:
                csvreader = csv.reader(csvfile)
                headers=next(csvreader)            
                for row in csvreader:
                    rows.append(row)    
            break
        except:
            print('Invalid Address')
    while True:
            col= input('Enter the name of the Column: ')      
            if col in headers:
                col_index= headers.index(col)
                print("processing")
                break
            else:
                print("invalid column")                
    email_list=[]
    for row in rows:
        email_list.append(row[col_index])
    not_valid_emails=[]
    for recv in email_list:          
        email_status=check(recv)
        if email_status=="Valid Email":
            msg['To'] = recv
            smtp_server.sendmail(Sender_Email,recv,msg.as_string())
        else:
            not_valid_emails.append(recv)
    print(f"successfully sent to {len(email_list)-len(not_valid_emails)} of {len(email_list)}")
    if len(not_valid_emails) == 0:
        pass
    else:
        print(f"not sent to {not_valid_emails}")
    smtp_server.quit()

print('''Welcome to the Bulk Gmail Sender!
NB: This is a third party application and by the access
of the third party applications are turned off in Google. 
So before we start you are requested to turn on the Less 
secured apps access of your gmail account! If your gmail
account is secured with two step verification, you might
need a specific password and it will become difficult for
us to access ^.^''')
smtp_server = smtplib.SMTP_SSL('smtp.gmail.com', 465)
while True:
    Sender_Email = input("From: ")
    email_status=check(Sender_Email)
    if email_status=="Valid Email":
        break
    else:
        print("Please enter a valid email ")
        pass
while True:
    Password = input('Password: ')
    try:
        smtp_server.login(Sender_Email,Password)
        print("Gmail authentication successful")
        break
    except:
        print("Wrong password or the third party applications are turned off in your Google Account")
        
msg = MIMEMultipart()
msg['Subject'] = input('Write the Subject: ')
msg['From'] = Sender_Email
msg_content = MIMEText(input('Write your mail here: '), 'plain')
msg.attach(msg_content)
print("""How do you want to input the mail addresses?
1. Adding Manually
2. Submit list
3. Uploading Csv file
4. Uploading Excel Sheet etc.
""")
while True:
    option=input("Enter 1,2,3 or 4: ")
    if option== "1":
        adding_manually()
        break
    elif option=="2":
        submit_list()
        break
    elif option=="3":
        csv_send()
        break
    elif option=="4":
        excel()
        break
    else:
        print("Unknown input")